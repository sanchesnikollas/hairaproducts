# tests/registry/test_excel_loader.py
import json
import pytest
from pathlib import Path
from unittest.mock import patch

from src.registry.excel_loader import load_brands_from_excel, Brand


# We'll mock openpyxl to avoid needing a real xlsx in tests
class TestLoadBrands:
    def test_loads_nacionais(self, tmp_path):
        # Create a minimal test xlsx
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Nacionais"
        ws.append(["Marca", "Marketplace", "Site da Marca", "Ingredientes no site"])
        ws.append(["Amend Cosméticos", "Beleza na Web", "https://www.amend.com.br/", "sim"])
        ws.append(["Test Brand", "Amazon", "https://www.testbrand.com.br/", "não"])
        filepath = tmp_path / "test.xlsx"
        wb.save(filepath)

        brands = load_brands_from_excel(str(filepath))
        assert len(brands) == 2
        assert brands[0].brand_name == "Amend Cosméticos"
        assert brands[0].brand_slug == "amend-cosmeticos"
        assert brands[0].official_url_root == "https://www.amend.com.br/"

    def test_loads_internacionais(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Nacionais"
        ws1.append(["Marca", "Marketplace", "Site da Marca", "Ingredientes no site"])

        ws2 = wb.create_sheet("Internacionais")
        ws2.append(["Marca", "País de Origem", "Marketplace", "Site da Marca", "Ingredientes no site da Marca"])
        ws2.append(["Kérastase", "França", "Sephora", "https://www.kerastase.com.br/", "sim"])
        filepath = tmp_path / "test.xlsx"
        wb.save(filepath)

        brands = load_brands_from_excel(str(filepath))
        assert len(brands) == 1
        assert brands[0].country == "França"

    def test_loads_marcas_principais_entrypoints(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Nacionais"
        ws1.append(["Marca", "Marketplace", "Site da Marca", "Ingredientes no site"])

        ws2 = wb.create_sheet("Marcas Principais")
        ws2.append(["Nome", "Site", "Caminho", "Extrair", "OBS"])
        ws2.append(["Wella", "https://loja.wella.com.br", "todo o site", None, None])
        ws2.append(["Truss", "https://www.trussprofessional.com.br", "Home > Produtos", "https://truss.com/produtos", None])
        filepath = tmp_path / "test.xlsx"
        wb.save(filepath)

        brands = load_brands_from_excel(str(filepath))
        wella = next(b for b in brands if b.brand_name == "Wella")
        assert "https://loja.wella.com.br" in wella.official_url_root
        assert wella.priority is not None  # Marcas Principais get priority

    def test_deduplicates_by_slug(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Nacionais"
        ws1.append(["Marca", "Marketplace", "Site da Marca", "Ingredientes no site"])
        ws1.append(["Amend", "BnW", "https://www.amend.com.br/", "sim"])

        ws2 = wb.create_sheet("Marcas Principais")
        ws2.append(["Nome", "Site", "Caminho", "Extrair", "OBS"])
        ws2.append(["Amend", "https://www.amend.com.br/", None, None, None])
        filepath = tmp_path / "test.xlsx"
        wb.save(filepath)

        brands = load_brands_from_excel(str(filepath))
        amend_brands = [b for b in brands if "amend" in b.brand_slug]
        assert len(amend_brands) == 1
        assert amend_brands[0].priority is not None  # Marcas Principais priority wins

    def test_export_to_json(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Nacionais"
        ws.append(["Marca", "Marketplace", "Site da Marca", "Ingredientes no site"])
        ws.append(["TestBrand", "BnW", "https://www.test.com.br/", "não"])
        filepath = tmp_path / "test.xlsx"
        wb.save(filepath)

        brands = load_brands_from_excel(str(filepath))
        output = tmp_path / "brands.json"
        from src.registry.excel_loader import export_brands_json
        export_brands_json(brands, str(output))

        data = json.loads(output.read_text())
        assert len(data) == 1
        assert data[0]["brand_name"] == "TestBrand"
