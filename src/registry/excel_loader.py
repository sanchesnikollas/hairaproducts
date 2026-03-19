# src/registry/excel_loader.py
from __future__ import annotations

import json
import logging
from pathlib import Path

import openpyxl
from slugify import slugify

from src.core.models import Brand

logger = logging.getLogger(__name__)


def _normalize_url(url: str) -> str:
    return url.strip().rstrip("/").lower()


def _clean_url(url: str | None) -> str:
    if not url:
        return ""
    return url.strip()


def _dedup_by_url(brands_by_slug: dict[str, Brand]) -> dict[str, Brand]:
    """Merge brands that share the same URL root, preferring the one with priority."""
    from collections import defaultdict

    url_groups: dict[str, list[str]] = defaultdict(list)
    for slug, brand in brands_by_slug.items():
        if brand.official_url_root:
            key = _normalize_url(brand.official_url_root)
            if key:
                url_groups[key].append(slug)

    for url_key, slugs in url_groups.items():
        if len(slugs) < 2:
            continue
        # Pick winner: prefer brand with priority, then first encountered
        winner_slug = slugs[0]
        for s in slugs:
            if brands_by_slug[s].priority is not None:
                winner_slug = s
                break
        winner = brands_by_slug[winner_slug]
        for s in slugs:
            if s == winner_slug:
                continue
            loser = brands_by_slug[s]
            # Absorb entrypoints
            for ep in loser.catalog_entrypoints:
                if ep not in winner.catalog_entrypoints:
                    winner.catalog_entrypoints.append(ep)
            # Absorb country if winner lacks it
            if not winner.country and loser.country:
                winner.country = loser.country
            # Absorb notes
            if loser.notes and not winner.notes:
                winner.notes = loser.notes
            del brands_by_slug[s]
            logger.debug(f"Merged duplicate brand {loser.brand_name!r} ({s}) into {winner.brand_name!r} ({winner_slug})")

    return brands_by_slug


def load_brands_from_excel(filepath: str) -> list[Brand]:
    wb = openpyxl.load_workbook(filepath, read_only=True)
    brands_by_slug: dict[str, Brand] = {}

    # 1. Load Nacionais
    if "Nacionais" in wb.sheetnames:
        ws = wb["Nacionais"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0] if row[0] else None
            site = row[2] if len(row) > 2 else None
            has_inci = row[3] if len(row) > 3 else None
            if not name:
                continue
            name = str(name).strip()
            slug = slugify(name)
            if not slug:
                continue
            brand = Brand(
                brand_name=name,
                brand_slug=slug,
                official_url_root=_clean_url(str(site)) if site else "",
                country="Brasil",
                notes=f"inci_on_site={has_inci}" if has_inci else None,
            )
            brands_by_slug[slug] = brand

    # 2. Load Internacionais
    if "Internacionais" in wb.sheetnames:
        ws = wb["Internacionais"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0] if row[0] else None
            country = row[1] if len(row) > 1 else None
            site = row[3] if len(row) > 3 else None
            if not name:
                continue
            name = str(name).strip()
            slug = slugify(name)
            if not slug:
                continue
            if slug not in brands_by_slug:
                brand = Brand(
                    brand_name=name,
                    brand_slug=slug,
                    official_url_root=_clean_url(str(site)) if site else "",
                    country=str(country).strip() if country else None,
                )
                brands_by_slug[slug] = brand

    # 3. Load Marcas Principais (these get priority + entrypoints)
    if "Marcas Principais" in wb.sheetnames:
        ws = wb["Marcas Principais"]
        current_brand_slug: str | None = None
        priority_counter = 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0] if row[0] else None
            site = row[1] if len(row) > 1 else None
            caminho = row[2] if len(row) > 2 else None
            extrair = row[3] if len(row) > 3 else None
            obs = row[4] if len(row) > 4 else None

            if name:
                name = str(name).strip()
                slug = slugify(name)
                current_brand_slug = slug
                if slug in brands_by_slug:
                    brands_by_slug[slug].priority = priority_counter
                    if site:
                        brands_by_slug[slug].official_url_root = _clean_url(str(site))
                    if obs:
                        brands_by_slug[slug].notes = str(obs).strip()
                else:
                    brand = Brand(
                        brand_name=name,
                        brand_slug=slug,
                        official_url_root=_clean_url(str(site)) if site else "",
                        priority=priority_counter,
                        notes=str(obs).strip() if obs else None,
                    )
                    brands_by_slug[slug] = brand
                priority_counter += 1

            # Add entrypoints from Caminho or Extrair columns
            target_slug = current_brand_slug
            if target_slug and target_slug in brands_by_slug:
                if extrair and str(extrair).startswith("http"):
                    url = _clean_url(str(extrair))
                    if url not in brands_by_slug[target_slug].catalog_entrypoints:
                        brands_by_slug[target_slug].catalog_entrypoints.append(url)
                elif site and str(site).startswith("http") and name:
                    url = _clean_url(str(site))
                    if url not in brands_by_slug[target_slug].catalog_entrypoints:
                        brands_by_slug[target_slug].catalog_entrypoints.append(url)

    wb.close()

    # 4. Dedup brands that share the same URL root but have different slugs
    brands_by_slug = _dedup_by_url(brands_by_slug)

    result = list(brands_by_slug.values())
    logger.info(f"Loaded {len(result)} brands from {filepath}")
    return result


def export_brands_json(brands: list[Brand], output_path: str) -> None:
    data = [b.model_dump() for b in brands]
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    logger.info(f"Exported {len(data)} brands to {output_path}")
